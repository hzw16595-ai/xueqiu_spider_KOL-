# -*- coding: utf-8 -*-
"""
从上证50语料库中抽样帖子用于人工情绪标注
- 每家企业至少2条，共50家企业
- 尽量均匀覆盖2023-2025年
- 不重复抽取（按post_id去重）
- 输出Excel
"""
import csv
import os
import re
import random
import pandas as pd
from collections import defaultdict

random.seed(42)

# 1. 读取上证50映射
mapping = pd.read_excel(r'D:\python-study\pythonProject3\企业股票代码映射.xlsx')
sz50_set = set(mapping['股票代码'].astype(str).str.zfill(6))
sz50_names = dict(zip(mapping['股票代码'].astype(str).str.zfill(6), mapping['匹配企业']))

# 2. 直接从全量CSV读取，筛选上证50，按企业+年份分组
all_file = r'C:\Users\HZW\xueqiu_spider\output\xueqiu_all_posts_merged_489542条.csv'
output_file = r'C:\Users\HZW\xueqiu_spider\output\xueqiu_sz50_sentiment_sample.xlsx'

ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
def clean(v):
    if not isinstance(v, str): return v
    return ILLEGAL_CHARS_RE.sub('', v)

company_year_posts = defaultdict(lambda: defaultdict(list))
seen_pids = set()
total = 0

print('Reading all posts & filtering SZ50...')
with open(all_file, 'rb') as f:
    raw = f.read()
# Remove BOM
if raw[:3] == b'\xef\xbb\xbf':
    raw = raw[3:]

text_io = io.BytesIO(raw) if 'io' in dir() else None
import io
text_io = io.BytesIO(raw)

reader = csv.DictReader(io.TextIOWrapper(text_io, encoding='utf-8', errors='replace'))
# Fix column names (may have issues)
for row in reader:
    total += 1
    # Get post_id - try different approaches
    pid = row.get('post_id', '') or ''
    if not pid:
        # First column value
        vals = list(row.values())
        if vals:
            pid = vals[0] or ''
    if not pid:
        continue
    if pid in seen_pids:
        continue
    seen_pids.add(pid)
    
    # Find matched SZ50 stocks
    found = {}
    ss = row.get('all_stocks', '') or ''
    if ss:
        for s in ss.replace(' ', '').split(';'):
            if not s: continue
            c = s[-6:] if len(s) >= 6 else s
            if c in sz50_set: found[c] = s
    text = row.get('text', '') or ''
    for code in sz50_set:
        if code in text and code not in found:
            found[code] = code
    
    if found:
        date = row.get('date', '') or ''
        year = date[:4] if len(date) >= 4 else 'unknown'
        for code in found:
            name = sz50_names.get(code, code)
            company_year_posts[name][year].append({
                '帖子ID': pid,
                '匹配企业': name,
                '股票代码': code,
                '发布时间': row.get('datetime', '') or '',
                '日期': date,
                '帖子正文': clean(text),
                'KOL名称': row.get('kol_name', '') or '',
                '点赞数': row.get('like_count', '') or '',
                '评论数': row.get('reply_count', '') or '',
                '转发数': row.get('retweet_count', '') or '',
            })
    
    if total % 100000 == 0:
        print(f'  {total:,} scanned', end='\r')

print(f'\nScanned: {total:,}, Unique posts with SZ50: {len(seen_pids)}')
print(f'Companies: {len(company_year_posts)}')

# 3. 抽样
sampled = []
sampled_pids = set()

# Step 1: 每家企业至少2条，覆盖不同年份
print('\nStep 1: 2 per company, different years...')
for company in sorted(company_year_posts.keys()):
    years = sorted(company_year_posts[company].keys())
    random.shuffle(years)
    picked = 0
    for year in years:
        posts = company_year_posts[company][year]
        avail = [p for p in posts if p['帖子ID'] not in sampled_pids]
        if avail and picked < 2:
            chosen = random.choice(avail)
            sampled.append(chosen)
            sampled_pids.add(chosen['帖子ID'])
            picked += 1
    while picked < 2:
        for year in sorted(company_year_posts[company].keys(), reverse=True):
            avail = [p for p in company_year_posts[company][year] if p['帖子ID'] not in sampled_pids]
            if avail:
                chosen = random.choice(avail)
                sampled.append(chosen)
                sampled_pids.add(chosen['帖子ID'])
                picked += 1
                break
        else:
            break

print(f'  After step 1: {len(sampled)}')

# Step 2: 补充到~110条
target = 110
while len(sampled) < target:
    company_counts = defaultdict(int)
    for s in sampled:
        company_counts[s['匹配企业']] += 1
    company_totals = {c: sum(len(ps) for ps in yp.values()) for c, yp in company_year_posts.items()}
    candidates = []
    for c in company_year_posts:
        if company_totals[c] > 0:
            candidates.append((company_counts.get(c, 0) / company_totals[c], c))
    candidates.sort()
    
    added = False
    for _, company in candidates:
        years = sorted(company_year_posts[company].keys())
        random.shuffle(years)
        for year in years:
            avail = [p for p in company_year_posts[company][year] if p['帖子ID'] not in sampled_pids]
            if avail:
                chosen = random.choice(avail)
                sampled.append(chosen)
                sampled_pids.add(chosen['帖子ID'])
                added = True
                break
        if added:
            break
    if not added:
        break

print(f'  After step 2: {len(sampled)}')

# 4. Stats
company_cnt = defaultdict(int)
year_cnt = defaultdict(int)
for s in sampled:
    company_cnt[s['匹配企业']] += 1
    d = s['日期']
    if d and len(d) >= 4:
        year_cnt[d[:4]] += 1

print(f'\n=== Sample Stats ===')
print(f'Total: {len(sampled)}')
print(f'Companies: {len(company_cnt)}/50')
print(f'Years: {dict(sorted(year_cnt.items()))}')

# 5. Write Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = 'sentiment_sample'

headers = ['No.', 'Post_ID', 'Company', 'Stock_Code', 'Datetime', 'Date',
           'Text', 'KOL', 'Likes', 'Replies', 'Retweets', 'Sentiment']
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
data_font = Font(size=10, name='Arial')
label_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Chinese header in row 1
cn_headers = ['序号', '帖子ID', '匹配企业', '股票代码', '发布时间', '日期',
              '帖子正文', 'KOL名称', '点赞数', '评论数', '转发数', '情绪标签']
for ci, (en, cn) in enumerate(zip(headers, cn_headers), 1):
    cell = ws.cell(row=1, column=ci, value=f'{en}\n{cn}')
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Sort by company + date
sampled.sort(key=lambda x: (x['匹配企业'], x['日期']))
row_map = {
    'No.': lambda s: None, 'Post_ID': lambda s: s['帖子ID'],
    'Company': lambda s: s['匹配企业'], 'Stock_Code': lambda s: s['股票代码'],
    'Datetime': lambda s: s['发布时间'], 'Date': lambda s: s['日期'],
    'Text': lambda s: s['帖子正文'], 'KOL': lambda s: s['KOL名称'],
    'Likes': lambda s: s['点赞数'], 'Replies': lambda s: s['评论数'],
    'Retweets': lambda s: s['转发数'], 'Sentiment': lambda s: '',
}

for ri, s in enumerate(sampled, 2):
    for ci, h in enumerate(headers, 1):
        val = ri - 1 if h == 'No.' else row_map[h](s)
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = data_font
        cell.border = thin_border
        if h == 'Sentiment':
            cell.fill = label_fill
        elif h == 'Text':
            cell.alignment = Alignment(vertical='top', wrap_text=True)

widths = [6, 12, 14, 12, 20, 12, 70, 14, 8, 8, 8, 14]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A2'
ws.row_dimensions[1].height = 30

# Instructions sheet
ws2 = wb.create_sheet('Labeling_Guide')
guide = [
    ['Sentiment Labeling Guide'],
    [''],
    ['Fill the "Sentiment" column with one of:'],
    [''],
    ['positive', 'Bullish on the matched stock/company, recommend buy, praise performance'],
    ['negative', 'Bearish, expect decline, recommend sell, criticize management'],
    ['neutral', 'Objective reporting, news relay, no clear sentiment'],
    [''],
    ['Notes:'],
    ['1. Judge sentiment TOWARD the matched company, not the whole market'],
    ['2. Focus only on the matched company if multiple stocks are mentioned'],
    ['3. Mark as neutral if content is irrelevant or sentiment is unclear'],
    ['4. For retweets, judge based on the original content'],
]
for ri, row_data in enumerate(guide, 1):
    for ci, v in enumerate(row_data, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = Font(bold=(ri==1), size=11, name='Arial')
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 70

wb.save(output_file)
sz = os.path.getsize(output_file)
print(f'\nOutput: {output_file} ({sz/1024:.0f}KB)')
print('Done!')
